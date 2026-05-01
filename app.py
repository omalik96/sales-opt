import io
import warnings

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from optimizer import ALL_COMBOS, DEFAULT_COMBOS, make_matrix, optimise

warnings.filterwarnings('ignore')

st.set_page_config(
    page_title='Transport Optimalizace',
    page_icon='🚛',
    layout='wide',
)

# ─── sidebar ──────────────────────────────────────────────────────────────────

with st.sidebar:
    st.title('🚛 Transport Optimalizace')
    st.divider()

    uploaded = st.file_uploader('Nahrát Excel soubor', type=['xlsx'])

    st.subheader('Kapacita vozidla')
    capacity = st.number_input(
        'Maximální počet palet', min_value=1, max_value=100, value=33, step=1,
        label_visibility='collapsed'
    )

    st.subheader('Povolené kombinace')
    combo_checks = {}
    for combo in ALL_COMBOS:
        combo_checks[combo] = st.checkbox(combo, value=True)
    allowed_combos = {c for c, v in combo_checks.items() if v}

    st.divider()
    run = st.button('🔄 Přepočítat', use_container_width=True, type='primary')

    st.divider()
    st.caption(
        '⚠️ **Upozornění:** Tato aplikace byla vytvořena metodou vibecoding '
        '(AI-assisted development). Uživatel je odpovědný za kontrolu a ověření '
        'všech výstupů před jejich dalším použitím.'
    )


# ─── data loading ─────────────────────────────────────────────────────────────

@st.cache_data
def load_data(source) -> tuple[pd.DataFrame, pd.DataFrame] | None:
    try:
        f = io.BytesIO(source) if source else 'Data claude.xlsx'
        df_ntw = pd.read_excel(f, sheet_name='NTW')
        df_115 = pd.read_excel(f, sheet_name='1 až 15')
        return df_ntw, df_115
    except Exception:
        return None


file_bytes = uploaded.read() if uploaded else None
result = load_data(file_bytes)

if result is None:
    st.info('👆 Nahraj Excel soubor v levém panelu pro spuštění optimalizace.')
    st.stop()

df_ntw, df_115 = result


# ─── computation ──────────────────────────────────────────────────────────────

# Auto-run on first load; re-run on button press
if 'results' not in st.session_state or run:
    with st.spinner('Probíhá optimalizace…'):
        trips_ntw, warn_ntw = optimise(df_ntw, allowed_combos, capacity)
        trips_115, warn_115 = optimise(df_115, allowed_combos, capacity)
        matrix_ntw = make_matrix(trips_ntw)
        matrix_115 = make_matrix(trips_115)
    st.session_state['results'] = (trips_ntw, trips_115, matrix_ntw, matrix_115, warn_ntw, warn_115)

trips_ntw, trips_115, matrix_ntw, matrix_115, warn_ntw, warn_115 = st.session_state['results']


# ─── helpers ──────────────────────────────────────────────────────────────────

def metrics_row(trips: pd.DataFrame, cap: int):
    total_p = int(trips['Palety'].sum()) if not trips.empty else 0
    avg_u = trips['Vytížení'].mean() * 100 if not trips.empty else 0
    full = int((trips['Palety'] == cap).sum()) if not trips.empty else 0
    c1, c2, c3, c4 = st.columns(4)
    c1.metric('Počet jízd', f'{len(trips):,}')
    c2.metric('Celkem palet', f'{total_p:,}')
    c3.metric('Průměrné vytížení', f'{avg_u:.1f} %')
    c4.metric(f'Plně naložených ({cap}p)', f'{full:,}')


def combo_chart(trips: pd.DataFrame):
    if trips.empty:
        return
    data = (
        trips.groupby('Kombinace dep')['Palety']
        .sum()
        .reset_index()
        .sort_values('Palety', ascending=False)
    )
    fig = px.bar(
        data, x='Kombinace dep', y='Palety',
        color='Kombinace dep',
        labels={'Palety': 'Celkem palet', 'Kombinace dep': 'Kombinace'},
        color_discrete_sequence=px.colors.qualitative.Safe,
    )
    fig.update_layout(showlegend=False, height=320, margin=dict(t=10, b=10))
    st.plotly_chart(fig, use_container_width=True)


def utilization_histogram(trips: pd.DataFrame, cap: int):
    if trips.empty:
        return
    fig = px.histogram(
        trips, x='Palety', nbins=cap,
        labels={'Palety': 'Počet palet na jízdu', 'count': 'Počet jízd'},
        color_discrete_sequence=['#1F4E79'],
    )
    fig.update_layout(height=280, margin=dict(t=10, b=10))
    st.plotly_chart(fig, use_container_width=True)


def trips_table(trips: pd.DataFrame):
    if trips.empty:
        st.info('Žádné jízdy – zkontroluj aktivní kombinace.')
        return
    display = trips.copy()
    display['Datum'] = display['Datum'].dt.strftime('%d.%m.%Y')
    display['Vytížení'] = display['Vytížení'].map('{:.1%}'.format)
    st.dataframe(display, use_container_width=True, hide_index=True, height=420)


def matrix_table(matrix: pd.DataFrame):
    if matrix.empty:
        st.info('Žádná data pro matici.')
        return
    st.dataframe(matrix, use_container_width=True, hide_index=True)


def show_warnings(warns: list):
    if warns:
        with st.expander(f'⚠️ Upozornění ({len(warns)})', expanded=False):
            for w in warns:
                st.warning(w)


# ─── export to Excel ──────────────────────────────────────────────────────────

HDR_FILL = PatternFill('solid', start_color='1F4E79')
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BODY_FONT = Font(name='Arial', size=10)
ALT_FILL = PatternFill('solid', start_color='D6E4F0')
_THIN = Side(style='thin', color='AAAAAA')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _fmt_sheet(ws, df: pd.DataFrame):
    ws.freeze_panes = 'A2'
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(1, ci, col)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _BORDER
    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            cell.font = BODY_FONT
            cell.border = _BORDER
            if fill:
                cell.fill = fill
            col_name = df.columns[ci - 1]
            if col_name == 'Vytížení':
                cell.number_format = '0.0%'
            elif col_name == 'Datum':
                cell.number_format = 'DD.MM.YYYY'
    for ci, col in enumerate(df.columns, 1):
        max_w = max(len(str(col)), *[len(str(ws.cell(r, ci).value or '')) for r in range(2, ws.max_row + 1)])
        ws.column_dimensions[get_column_letter(ci)].width = min(max_w + 2, 45)
    ws.row_dimensions[1].height = 30


def _fmt_matrix(ws, mdf: pd.DataFrame):
    ws['A1'] = 'Matice palet: kombinace × měsíc'
    ws['A1'].font = Font(name='Arial', bold=True, size=11)
    for ci, col in enumerate(mdf.columns, 1):
        cell = ws.cell(2, ci, col)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = _BORDER
    for ri, row in enumerate(mdf.itertuples(index=False), 3):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            cell.font = BODY_FONT
            cell.border = _BORDER
            if ci > 1:
                cell.alignment = Alignment(horizontal='center')
    for ci in range(1, len(mdf.columns) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 12
    ws.column_dimensions['A'].width = 22
    ws.freeze_panes = 'B3'


def build_excel(trips_ntw, trips_115, matrix_ntw, matrix_115) -> bytes:
    try:
        wb = load_workbook('Data claude.xlsx')
    except Exception:
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)

    def add_or_replace(name):
        if name in wb.sheetnames:
            del wb[name]
        return wb.create_sheet(name)

    _fmt_sheet(add_or_replace('NTW – jízdy (opt)'), trips_ntw)
    _fmt_matrix(add_or_replace('NTW – matice (opt)'), matrix_ntw)
    _fmt_sheet(add_or_replace('1až15 – jízdy (opt)'), trips_115)
    _fmt_matrix(add_or_replace('1až15 – matice (opt)'), matrix_115)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── main tabs ────────────────────────────────────────────────────────────────

tab_ntw, tab_115, tab_sum = st.tabs(['NTW', '1 až 15', 'Souhrn'])

with tab_ntw:
    st.subheader('NTW – navržené jízdy')
    metrics_row(trips_ntw, capacity)
    show_warnings(warn_ntw)

    col_chart, col_hist = st.columns(2)
    with col_chart:
        st.caption('Palety dle kombinace dep')
        combo_chart(trips_ntw)
    with col_hist:
        st.caption('Histogram vytížení vozidel')
        utilization_histogram(trips_ntw, capacity)

    st.caption('Detail jízd')
    trips_table(trips_ntw)

    st.caption('Matice palet: kombinace × měsíc')
    matrix_table(matrix_ntw)

with tab_115:
    st.subheader('1 až 15 – navržené jízdy')
    metrics_row(trips_115, capacity)
    show_warnings(warn_115)

    col_chart, col_hist = st.columns(2)
    with col_chart:
        st.caption('Palety dle kombinace dep')
        combo_chart(trips_115)
    with col_hist:
        st.caption('Histogram vytížení vozidel')
        utilization_histogram(trips_115, capacity)

    st.caption('Detail jízd')
    trips_table(trips_115)

    st.caption('Matice palet: kombinace × měsíc')
    matrix_table(matrix_115)

with tab_sum:
    st.subheader('Souhrn optimalizace')

    def summary_df(trips, label):
        if trips.empty:
            return {}
        return {
            'List': label,
            'Počet jízd': len(trips),
            'Celkem palet': int(trips['Palety'].sum()),
            'Průměrné vytížení': f"{trips['Vytížení'].mean()*100:.1f} %",
            f'Plně naložených ({capacity}p)': int((trips['Palety'] == capacity).sum()),
            'Nejčastější kombinace': trips['Kombinace dep'].value_counts().index[0],
        }

    rows = [summary_df(trips_ntw, 'NTW'), summary_df(trips_115, '1 až 15')]
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    st.divider()

    col_ntw, col_115 = st.columns(2)
    with col_ntw:
        st.caption('NTW – jízdy dle kombinace a měsíce')
        if not trips_ntw.empty:
            pivot = trips_ntw.pivot_table(index='Kombinace dep', columns='Měsíc', values='Č. jízdy', aggfunc='count', fill_value=0)
            st.dataframe(pivot, use_container_width=True)
    with col_115:
        st.caption('1 až 15 – jízdy dle kombinace a měsíce')
        if not trips_115.empty:
            pivot = trips_115.pivot_table(index='Kombinace dep', columns='Měsíc', values='Č. jízdy', aggfunc='count', fill_value=0)
            st.dataframe(pivot, use_container_width=True)

    st.divider()
    st.subheader('Export výsledků')
    excel_bytes = build_excel(trips_ntw, trips_115, matrix_ntw, matrix_115)
    st.download_button(
        label='⬇️ Stáhnout výsledky jako Excel',
        data=excel_bytes,
        file_name='optimalizace_výsledky.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        use_container_width=True,
    )
